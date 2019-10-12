'''import openpyxl

wb=openpyxl.Workbook() 
sheet = wb.active
sheet.title = 'new title'
sheet[A1] = '男人'
row = ['美国队长','蜘蛛侠','钢铁侠']
sheet.append(row)
rows = [['美国队长','钢铁侠','蜘蛛侠'],['是','漫威','宇宙', '经典','人物']]
#先把要写入的多行内容写成列表，再放进大列表里，赋值给rows。
for i in rows:
    sheet.append(i)
#遍历rows，同时把遍历的内容添加到表格里，这样就实现了多行写入。
print(rows)
wb.save('Marvel.xlsx')'''
import openpyxl 
wb=openpyxl.Workbook() 
sheet=wb.active
sheet.title='new title'
sheet['A1'] = '漫威宇宙'
rows= [['美国队长','钢铁侠','蜘蛛侠'],['是','漫威','宇宙', '经典','人物']]
for i in rows:
    sheet.append(i)
wb.save('Marvel.xlsx')

wb = openpyxl.load_workbook('Marvel.xlsx')
sheet = wb['new title']
sheetname = wb.sheetnames
print(sheetname)
A1_value = sheet['A1'].value
print(A1_value)

wb = openpyxl.load_workbook('Marvel.xlsx')
sheet = wb['new title']
sheetname = wb.sheetnames
print(sheetname)
A1_value = sheet['A1'].value
print(A1_value)